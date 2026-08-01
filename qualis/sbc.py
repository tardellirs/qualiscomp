"""Planilha oficial de conferências das Comissões Especiais da SBC.

O Documento de Área não lista os eventos: ele remete às indicações das CEs da
SBC (nota de rodapé 8, apontando para sbc.org.br/documentosinstitucionais).
A planilha vigente é:

    https://www.sbc.org.br/wp-content/uploads/2024/11/
    CEs-SBC-Conferências-versão-de-2024.xlsx

São 30 abas (uma por Comissão Especial), cada uma classificando eventos como
"Top 10", "Top 20" ou "Relevante" — exatamente os rótulos que a Área 02 usa para
reclassificar artigos (+2, +1, mantém).

Duas armadilhas conhecidas:

1. A URL só resolve com o nome de arquivo em **NFD** (decomposto). Em NFC dá 404.
2. A coluna H5 da planilha é uma coleta da SBC de 2024 e **diverge** do Google
   Scholar atual (SBES 21 na planilha vs 23 hoje; LADC 14 vs 7). A autoridade do
   h5 é o Scholar (`qualis.scholar`); a planilha vale pelo rótulo Top10/Top20.
"""

from __future__ import annotations

import unicodedata
import urllib.parse
import urllib.request
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

BASE = "https://www.sbc.org.br/wp-content/uploads/2024/11/"
ARQUIVO = "CEs-SBC-Conferências-versão-de-2024.xlsx"
CACHE = Path(__file__).resolve().parent.parent / "data" / "ces_sbc.xlsx"

# Abas que não são de Comissão Especial.
_ABAS_IGNORADAS = {"Instruções", "NomesPadrão"}

# Ordem de preferência quando o mesmo evento é classificado por várias CEs.
_PRIORIDADE = {"top10": 0, "top20": 1, "relevante": 2}


@dataclass
class EventoSBC:
    sigla: str
    nome: str
    classificacao: str  # top10 | top20 | relevante
    h5_sbc: int | None
    ces: list[str]
    nomes_alternativos: list[str] = field(default_factory=list)

    @property
    def todos_os_nomes(self) -> list[str]:
        vistos, out = set(), []
        for n in [self.nome, *self.nomes_alternativos]:
            if n and n.lower() not in vistos:
                vistos.add(n.lower())
                out.append(n)
        return out

    @property
    def multiplas_ces(self) -> bool:
        return len(self.ces) > 1


def baixar(destino: Path | None = None, *, forcar: bool = False) -> Path:
    """Baixa a planilha da SBC (usando NFD no nome, senão dá 404)."""
    destino = destino or CACHE
    if destino.exists() and not forcar:
        return destino
    url = BASE + urllib.parse.quote(unicodedata.normalize("NFD", ARQUIVO))
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=120) as r:
        dados = r.read()
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_bytes(dados)
    return destino


def _normalizar_classificacao(v: object) -> str | None:
    """Mapeia o rótulo da planilha para o vocabulário do Documento de Área.

    A planilha não usa a palavra "relevante": a terceira faixa se chama
    **"Eventos da Área"** (1868 linhas), que é o que o documento chama de
    eventos "indicados como relevantes para as CE-SBC". As faixas Top vêm
    grafadas de sete jeitos diferentes ("Top 10", "TOP10", "TOP-20"...).

    Dois rótulos não existem no documento: "Top, mas não é core" e "Top, mas é
    jovem". São ressalvas da própria SBC. Tratamos como `relevante` (mantém a
    classificação do h5) em vez de conceder +2 níveis: conceder um bônus que a
    SBC explicitamente relativizou seria inflar estrato sem respaldo.
    """
    t = str(v or "").strip().lower().replace("-", " ")
    t_junto = t.replace(" ", "")
    if t_junto in ("top", "sigla", ""):
        return None  # cabeçalhos
    if t_junto.startswith("top,mas") or t.startswith("top, mas"):
        return "relevante"
    if t_junto.startswith("top10"):
        return "top10"
    if t_junto.startswith("top20"):
        return "top20"
    if t_junto.startswith("relevante") or t_junto.startswith("eventosdaárea"):
        return "relevante"
    if t_junto.startswith("eventosdaarea"):
        return "relevante"
    return None


ALIASES = Path(__file__).resolve().parent.parent / "data" / "aliases_eventos.csv"
RECUSADOS = Path(__file__).resolve().parent.parent / "data" / "apelidos_recusados.csv"
NOMES_CE = Path(__file__).resolve().parent.parent / "data" / "ces_sbc_nomes.csv"


def nomes_das_ces(caminho: Path | None = None) -> dict[str, str]:
    """Sigla da Comissão Especial -> nome por extenso.

    A planilha de conferências traz só a sigla, no nome da aba. Sem o nome,
    "CE-GRAPI" não diz nada a quem não é da subárea.
    """
    import csv

    caminho = caminho or NOMES_CE
    if not caminho.exists():
        return {}
    with caminho.open(encoding="utf-8") as f:
        linhas = [l for l in f if not l.lstrip().startswith("//")]
    return {
        (r["sigla"] or "").strip(): (r["nome"] or "").strip()
        for r in csv.DictReader(linhas)
        if (r.get("sigla") or "").strip()
    }


def apelidos_recusados(caminho: Path | None = None) -> set[tuple[str, str]]:
    """(sigla, apelido em minúsculas) que a planilha da SBC atribui ao evento
    errado. Ver o cabeçalho do CSV."""
    import csv

    caminho = caminho or RECUSADOS
    if not caminho.exists():
        return set()
    with caminho.open(encoding="utf-8") as f:
        linhas = [ln for ln in f if not ln.lstrip().startswith("//")]
    return {
        ((r.get("sigla") or "").strip().upper(), (r.get("alias") or "").strip().lower())
        for r in csv.DictReader(linhas)
        if (r.get("sigla") or "").strip() and (r.get("alias") or "").strip()
    }


def carregar_aliases(caminho: Path | None = None) -> dict[str, list[str]]:
    """Apelidos curados à mão, por sigla (ver o cabeçalho do CSV)."""
    import csv

    caminho = caminho or ALIASES
    if not caminho.exists():
        return {}
    with caminho.open(encoding="utf-8") as f:
        linhas = [l for l in f if not l.lstrip().startswith("//")]
    out: dict[str, list[str]] = {}
    for row in csv.DictReader(linhas):
        sigla = (row.get("sigla") or "").strip().upper()
        alias = (row.get("alias") or "").strip()
        if sigla and alias:
            out.setdefault(sigla, []).append(alias)
    return out


ABA_EVENTOS_SBC = "Eventos SBC"
TRADICAO = Path(__file__).resolve().parent.parent / "data" / "tradicao_eventos.csv"


@dataclass(frozen=True)
class EventoDaSBC:
    """Evento promovido pela SBC, com o que mede sua tradição.

    A aba "Eventos SBC" da planilha é o que habilita o **critério de indução**
    do Documento de Área: "artigos publicados em eventos com pelo menos 20 anos
    de tradição poderão ser classificados no nível A4 e [...] pelo menos 10 anos
    [...] no nível A5". Sem ela, o CSBC — que existe desde 1980 — saía como A8.
    """

    sigla: str
    nome: str
    edicoes: int | None
    ano_primeira: int | None
    ce: str

    @property
    def anos_de_tradicao(self) -> int | None:
        """Medida conservadora: o número de EDIÇÕES.

        A planilha traz o ano da primeira edição em só 20 dos 88 eventos, mas o
        número de edições em todos. Para evento anual, edições <= anos desde a
        primeira (o SBES tem 37 edições e existe desde 1987). Usar edições nunca
        superestima a tradição — e superestimar aqui inflaria estrato.
        """
        return self.edicoes


def eventos_da_sbc(caminho: Path | None = None) -> dict[str, EventoDaSBC]:
    """Mapa sigla (maiúscula) -> evento promovido pela SBC."""
    from openpyxl import load_workbook

    caminho = caminho or baixar()
    wb = load_workbook(caminho, read_only=True, data_only=True)
    if ABA_EVENTOS_SBC not in wb.sheetnames:
        return {}

    def _int(v: object) -> int | None:
        t = str(v or "").strip()
        return int(t) if t.isdigit() else None

    out: dict[str, EventoDaSBC] = {}
    linhas = wb[ABA_EVENTOS_SBC].iter_rows(values_only=True)
    next(linhas, None)  # Sigla | Nome | SOL | edições | ano | CE | ...
    for r in linhas:
        if not r or not r[0]:
            continue
        sigla = str(r[0]).strip().rstrip("*").strip()
        if not sigla or sigla.lower() == "sigla":
            continue
        out[sigla.upper()] = EventoDaSBC(
            sigla=sigla,
            nome=str(r[1] or "").strip(),
            edicoes=_int(r[3]) if len(r) > 3 else None,
            ano_primeira=_int(r[4]) if len(r) > 4 else None,
            ce=str(r[5] or "").strip() if len(r) > 5 else "",
        )

    # Complemento curado, com fonte, para o que a aba não cobre.
    from datetime import date

    import csv as _csv

    if TRADICAO.exists():
        with TRADICAO.open(encoding="utf-8") as f:
            crus = [l for l in f if not l.lstrip().startswith("//")]
        for row in _csv.DictReader(crus):
            sigla = (row.get("sigla") or "").strip()
            desde = (row.get("desde") or "").strip()
            if not sigla or not desde.isdigit() or sigla.upper() in out:
                continue
            out[sigla.upper()] = EventoDaSBC(
                sigla=sigla,
                nome="",
                edicoes=date.today().year - int(desde),
                ano_primeira=int(desde),
                ce="",
            )
    return out


def ler(caminho: Path | None = None) -> list[EventoSBC]:
    """Lê todas as abas de CE e devolve os eventos deduplicados por sigla.

    Um mesmo evento aparece em várias CEs, às vezes com rótulos diferentes (o
    SBRC é Top20 na CE-RESD e "relevante" na CE-TF). Mantemos a MELHOR
    classificação e registramos todas as CEs em `ces`, porque o documento não
    diz como desempatar — quem decide é a comissão, e você precisa ver que houve
    divergência.
    """
    from openpyxl import load_workbook

    caminho = caminho or baixar()
    wb = load_workbook(caminho, read_only=True, data_only=True)

    por_sigla: dict[str, EventoSBC] = {}
    ces_por_sigla: dict[str, set[str]] = defaultdict(set)

    for ws in wb.worksheets:
        if ws.title in _ABAS_IGNORADAS:
            continue
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 3:
                continue
            classificacao = _normalizar_classificacao(row[0])
            if classificacao is None:
                continue
            # O asterisco é marcador de nota de rodapé da planilha, não parte
            # da sigla — sem tirar, "SBGames" e "SBGames*" viram dois eventos.
            sigla = str(row[1] or "").strip().rstrip("*").strip()
            nome = str(row[2] or "").strip()
            if not sigla or sigla.upper() == "SIGLA":
                continue
            # A planilha tem linhas de interface (célula de menu suspenso) que
            # não são eventos: sigla "escolher" e nome vazio.
            if not nome or sigla.lower() in ("escolher", "selecione", "nome"):
                continue
            h5 = row[3] if len(row) > 3 else None
            h5 = int(h5) if isinstance(h5, (int, float)) else None
            # Colunas "Novo Nome" (6) e "Nome Alternativo" (7): variantes do
            # mesmo nome, úteis para achar a entrada certa no Scholar.
            alternativos = [
                str(row[i]).strip()
                for i in (6, 7)
                if len(row) > i and row[i] and str(row[i]).strip()
            ]

            chave = sigla.upper()
            ces_por_sigla[chave].add(ws.title)
            atual = por_sigla.get(chave)
            if atual is None or (
                _PRIORIDADE[classificacao] < _PRIORIDADE[atual.classificacao]
            ):
                por_sigla[chave] = EventoSBC(
                    sigla, nome, classificacao, h5, [], alternativos
                )
            else:
                for a in alternativos:
                    if a not in atual.nomes_alternativos:
                        atual.nomes_alternativos.append(a)

    aliases = carregar_aliases()
    recusados = apelidos_recusados()
    for chave, ev in por_sigla.items():
        ev.nomes_alternativos = [
            a for a in ev.nomes_alternativos
            if (chave, a.strip().lower()) not in recusados
        ]
        ev.ces = sorted(ces_por_sigla[chave])
        for a in aliases.get(chave, []):
            if a not in ev.nomes_alternativos:
                ev.nomes_alternativos.append(a)
    return sorted(por_sigla.values(), key=lambda e: e.sigla)
