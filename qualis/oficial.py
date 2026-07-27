"""Qualis Eventos OFICIAL da CAPES para a área de Computação.

Baixado da consulta pública da Plataforma Sucupira (legado):

    https://sucupira-legado.capes.gov.br/sucupira/public/consultas/coleta/
    qualisEventos/listaQualisEventos.xhtml
    (form:evento=202 "Classificação de trabalho em anais 2025", form:area=2)

São 781 eventos classificados pela comissão da área. **É do quadriênio
2021-2024**, o ciclo encerrado — não substitui o cálculo para 2025-2028. Mas
vale muito por duas razões:

1. É a única classificação oficial de eventos que existe, feita pela mesma
   comissão e pelo mesmo método (h5 do Google + análise qualitativa das CEs da
   SBC), com cortes de h5 idênticos aos do ciclo novo.
2. Permitiu **decidir empiricamente** a ambiguidade da saturação em A3 (veja
   `test_saturacao_oficial.py`): entre os eventos Top10/Top20 cujo h5 sozinho
   daria pior que A3, só 3 de 93 foram classificados acima de A3 — enquanto no
   grupo de controle, cujo h5 já justificava A1/A2, 109 de 140 ficaram acima.
   Ou seja: o teto limita o ganho qualitativo, não o h5.

A escala oficial usa A1-A4 e B1-B4. São os mesmos 8 estratos do ciclo novo, com
os mesmos cortes de h5 (A1≥35 … B4>0), então B1..B4 correspondem a A5..A8.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

ARQUIVO = (
    Path(__file__).resolve().parent.parent
    / "data"
    / "qualis_eventos_oficial_2021_2024.xlsx"
)

# A escala antiga tem os mesmos 8 degraus e os mesmos cortes de h5.
ESCALA = {
    "A1": "A1", "A2": "A2", "A3": "A3", "A4": "A4",
    "B1": "A5", "B2": "A6", "B3": "A7", "B4": "A8",
}

CICLO = "2021-2024"


@dataclass(frozen=True)
class EventoOficial:
    sigla: str
    nome: str
    estrato_original: str  # como a CAPES publicou (A1..A4, B1..B4)
    estrato: str  # convertido para a escala A1..A8 do ciclo novo


def carregar(caminho: Path | None = None) -> dict[str, EventoOficial]:
    """Mapa sigla (maiúscula) -> classificação oficial."""
    from openpyxl import load_workbook

    caminho = caminho or ARQUIVO
    if not caminho.exists():
        return {}

    wb = load_workbook(caminho, read_only=True, data_only=True)
    linhas = wb.worksheets[0].iter_rows(values_only=True)
    next(linhas, None)  # cabeçalho: Sigla | Nome do evento | Estrato

    out: dict[str, EventoOficial] = {}
    for row in linhas:
        if not row or not row[0] or len(row) < 3 or not row[2]:
            continue
        sigla = str(row[0]).strip()
        bruto = str(row[2]).strip().upper()
        if bruto not in ESCALA:
            continue
        out[sigla.upper()] = EventoOficial(
            sigla=sigla,
            nome=str(row[1] or "").strip(),
            estrato_original=bruto,
            estrato=ESCALA[bruto],
        )
    return out
