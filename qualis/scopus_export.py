"""Importa exports da tela Scopus Sources (scopus.com/sources.uri).

Esse export é a MELHOR fonte disponível para a regra da Área 02, porque a coluna
"Highest percentile" já é o maior percentil do periódico entre suas categorias —
exatamente o que o documento pede ("o maior entre os dois", e o maior entre as
categorias). Formato da célula:

    99.0%
    1/1022
    Computer Science Applications

Vários arquivos podem ser importados juntos (a tela do Scopus limita o export);
em caso de duplicata, fica o maior percentil.
"""

from __future__ import annotations

import gzip
import json
import re
from dataclasses import asdict, dataclass, field, fields
from pathlib import Path

BD = Path(__file__).resolve().parent.parent / "data" / "scopus_percentis.json.gz"

_PCT = re.compile(r"([\d.]+)\s*%")
_RANK = re.compile(r"(\d+)\s*/\s*(\d+)")

# Periódicos cujo TÍTULO parece de evento. São revistas de verdade, com
# percentil Scopus, e várias são de altíssimo nível em Computação — o PACMPL
# é onde saem POPL, OOPSLA, ICFP e PLDI. Bloqueá-las como "evento" tornaria
# impossível classificar justamente os veículos mais fortes da área.
_PERIODICOS_COM_NOME_DE_EVENTO = (
    "proceedings of the acm on",
    "proceedings of the ieee",
    "proceedings of the national academy",
    "proceedings of the royal society",
    "proceedings of the vldb endowment",
    "proceedings of machine learning research",
)


@dataclass
class Fonte:
    """Um periódico (ou evento indexado) com seu percentil Scopus."""

    titulo: str
    percentil: float
    citescore: float | None
    rank: int | None
    total: int | None
    categoria: str | None
    snip: float | None
    sjr: float | None
    editora: str | None
    # Vem só da API da Elsevier; o export da tela Scopus Sources não traz ISSN.
    issns: list[str] = field(default_factory=list)
    ano_citescore: int | None = None
    # Áreas de topo do ASJC da revista (COMP, MULT, MEDI...). Vem da API; o
    # export manual não traz, e aí inferimos pelo nome da categoria.
    areas: list[str] = field(default_factory=list)

    @property
    def e_computacao(self) -> bool:
        """A revista pertence à área de Computação do ASJC.

        A regra da CAPES não exige isso — artigo de Computação em revista de
        outra área conta igual. Serve só para ordenar a tela de abertura por
        quem chega: gente da Computação.
        """
        if self.areas:
            return "COMP" in self.areas
        c = (self.categoria or "").lower()
        return any(
            m in c
            for m in (
                "comput", "software", "information system", "artificial",
                "vision and pattern", "hardware", "networks and communi",
                "signal processing", "human-computer", "theoretical comput",
            )
        )

    @property
    def e_sbc(self) -> bool:
        """Periódico da SBC — elegível a subir até 2 níveis pela análise da área.

        O Scopus grafa a editora em inglês ("Brazilian Computing Society"), não
        em português, então checar só "Sociedade Brasileira de Computação"
        deixaria o JBCS de fora justamente do bônus que mais lhe importa.
        """
        alvo = (self.editora or "").lower()
        return any(
            m in alvo
            for m in (
                "sociedade brasileira de computa",
                "brazilian computing society",
                "brazilian computer society",
            )
        )

    @property
    def parece_evento(self) -> bool:
        """True quando o título indica anais de evento, não periódico.

        Importa porque o export do Scopus mistura os dois, e a Área 02 usa
        regras diferentes: periódico vai por percentil, evento vai por h5 do
        Google Scholar. Aplicar percentil a um evento é usar a regra errada.

        Heurística, não verdade: existem PERIÓDICOS com "Proceedings" no nome
        (veja `_PERIODICOS_COM_NOME_DE_EVENTO`). Por isso quem chama deve
        avisar, não bloquear.
        """
        t = self.titulo.lower()
        if any(exc in t for exc in _PERIODICOS_COM_NOME_DE_EVENTO):
            return False
        return any(
            m in t
            for m in (
                "conference",
                "proceedings",
                "symposium",
                "workshop",
                "congress",
                "lecture notes",
                "annual meeting",
                "colloquium",
            )
        )


def normalizar(titulo: str) -> str:
    """Chave de comparação: minúsculas, sem pontuação, espaços colapsados."""
    t = titulo.lower().strip()
    t = re.sub(r"[^\w\s]", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def _float(v: object) -> float | None:
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _parse_percentil(celula: object) -> tuple[float | None, int | None, int | None, str | None]:
    """Decompõe a célula 'Highest percentile' em (percentil, rank, total, categoria)."""
    txt = str(celula or "").strip()
    if not txt or txt.lower() == "none":
        return None, None, None, None
    pct = _PCT.search(txt)
    rank = _RANK.search(txt)
    linhas = [l.strip() for l in txt.splitlines() if l.strip()]
    categoria = linhas[-1] if len(linhas) >= 3 else None
    return (
        float(pct.group(1)) if pct else None,
        int(rank.group(1)) if rank else None,
        int(rank.group(2)) if rank else None,
        categoria,
    )


def ler_export(caminho: str | Path) -> list[Fonte]:
    """Lê um export .xlsx da tela Scopus Sources."""
    from openpyxl import load_workbook

    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = [str(h or "").strip().lower() for h in next(rows)]

    def col(*nomes: str) -> int | None:
        for n in nomes:
            for i, h in enumerate(header):
                if h.startswith(n):
                    return i
        return None

    i_tit = col("source title", "title")
    i_pct = col("highest percentile", "percentile")
    if i_tit is None or i_pct is None:
        raise ValueError(
            f"{caminho}: não achei as colunas 'Source title' e 'Highest percentile'. "
            f"Cabeçalho lido: {header}"
        )
    i_cs, i_snip, i_sjr, i_ed = col("citescore"), col("snip"), col("sjr"), col("publisher")

    out: list[Fonte] = []
    for r in rows:
        if not r or i_tit >= len(r):
            continue
        titulo = str(r[i_tit] or "").strip()
        if not titulo:
            continue
        pct, rank, total, categoria = _parse_percentil(r[i_pct] if i_pct < len(r) else None)
        if pct is None:
            continue
        out.append(
            Fonte(
                titulo=titulo,
                percentil=pct,
                citescore=_float(r[i_cs]) if i_cs is not None and i_cs < len(r) else None,
                rank=rank,
                total=total,
                categoria=categoria,
                snip=_float(r[i_snip]) if i_snip is not None and i_snip < len(r) else None,
                sjr=_float(r[i_sjr]) if i_sjr is not None and i_sjr < len(r) else None,
                editora=(
                    str(r[i_ed]).strip() if i_ed is not None and i_ed < len(r) and r[i_ed] else None
                ),
            )
        )
    return out


def importar(caminhos: list[str | Path], *, mesclar: bool = True) -> dict[str, Fonte]:
    """Importa um ou mais exports e grava a base local.

    Duplicatas (mesmo título normalizado) ficam com o maior percentil, que é a
    convenção da própria regra.
    """
    bd = carregar() if mesclar and BD.exists() else {}
    novos = duplicatas = 0
    for c in caminhos:
        for f in ler_export(c):
            k = normalizar(f.titulo)
            atual = bd.get(k)
            if atual is None:
                bd[k] = f
                novos += 1
            else:
                duplicatas += 1
                if f.percentil > atual.percentil:
                    bd[k] = f
    BD.parent.mkdir(parents=True, exist_ok=True)
    with gzip.open(BD, "wt", encoding="utf-8") as fh:
        json.dump({k: asdict(v) for k, v in bd.items()}, fh, ensure_ascii=False)
    print(f"Importados {novos} novos, {duplicatas} duplicatas resolvidas pelo maior percentil.")
    print(f"Base local: {BD} ({len(bd)} fontes)")
    return bd


def carregar() -> dict[str, Fonte]:
    if not BD.exists():
        return {}
    with gzip.open(BD, "rt", encoding="utf-8") as fh:
        campos = {f.name for f in fields(Fonte)}
        return {
            k: Fonte(**{c: v for c, v in d.items() if c in campos})
            for k, d in json.load(fh).items()
        }


# Palavras genéricas demais para sustentar um match sozinhas.
_STOP = frozenset(
    "of the and for on in a an journal international transactions proceedings "
    "conference computer computing computers science sciences research letters "
    "review reviews advances".split()
)


def buscar(consulta: str, bd: dict[str, Fonte] | None = None, limite: int = 8) -> list[Fonte]:
    """Busca por título: exato, depois prefixo/substring, depois tokens.

    Nunca casa pelo caminho inverso (título contido na consulta): era isso que
    fazia "Journal of the Brazilian Computer Society" casar com a revista
    "Computer".
    """
    bd = bd if bd is not None else carregar()
    q = normalizar(consulta)
    if not q:
        return []
    if q in bd:
        return [bd[q]]

    # A consulta precisa estar contida no título, não o contrário.
    parciais = [f for k, f in bd.items() if q in k]
    if parciais:
        parciais.sort(key=lambda f: (len(f.titulo), -f.percentil))
        return parciais[:limite]

    # Último recurso: todos os tokens significativos da consulta no título.
    tokens = {t for t in q.split() if t not in _STOP}
    if not tokens:
        return []
    candidatos = [(k, f) for k, f in bd.items() if tokens <= set(k.split())]
    candidatos.sort(key=lambda kf: (len(kf[1].titulo), -kf[1].percentil))
    return [f for _, f in candidatos[:limite]]
