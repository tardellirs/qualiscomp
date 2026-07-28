"""Qualis Periódicos oficial da Computação, ciclo a ciclo (2010-2024).

Antes desta série a ficha mostrava o percentil Scopus ano a ano. Era um dado
correto e inútil: para revista pouco indexada a série parava em 2019, e um
histórico que termina há sete anos faz o site parecer abandonado — além de não
responder o que a pessoa pergunta, que é "como a CAPES via esta revista?".

Aqui está a resposta, e ela é oficial: as planilhas "Classificações publicadas
— todas as áreas de avaliação" da Plataforma Sucupira, filtradas pela área da
Computação. Quatro ciclos, cada um com o estrato que a comissão de fato
publicou.

**A escala mudou no meio do caminho**, e por isso os estratos NÃO são
comparáveis entre ciclos nem convertidos para a escala nova:

    2010-2012, 2013-2016   A1 A2 B1 B2 B3 B4 B5 C
    2017-2020, 2021-2024   A1 A2 A3 A4 B1 B2 B3 B4 C
    2025-2028              A1 ... A8          (o que este site estima)

Um "A2" de 2013 e um "A2" de 2021 saíram de réguas diferentes. Mostramos cada
ciclo com o rótulo que ele teve, sem traduzir — traduzir seria inventar uma
equivalência que a CAPES nunca publicou.

O casamento é por **ISSN**. Casar por título acrescentaria 5 revistas em 2.692
e traria erro grosso junto: existem duas revistas chamadas "Internet of Things",
uma C e outra A1.

A área muda de nome entre ciclos ("CIÊNCIA DA COMPUTAÇÃO" até 2020,
"COMPUTAÇÃO" em 2021-2024), então filtramos por substring.

As planilhas ficam em `data/` e fora do versionamento, como as demais fontes de
terceiros. Baixe em sucupira.capes.gov.br (Qualis > Classificações publicadas).
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

DADOS = Path(__file__).resolve().parent.parent / "data"
PADRAO = "classifica*_*-*.xls*"

# Ordem cronológica; a ficha mostra do mais antigo ao mais recente.
CICLOS = ("2010-2012", "2013-2016", "2017-2020", "2021-2024")


def normalizar_issn(v: object) -> str:
    return re.sub(r"[^0-9Xx]", "", str(v or "")).upper()


def _linhas(caminho: Path) -> list[tuple]:
    """ISSN | Título | Área de Avaliação | Estrato.

    Os arquivos com extensão `.xls` dos ciclos antigos não são XLS de verdade:
    são TSV com aspas. Tentar abrir como planilha binária falha.
    """
    if caminho.suffix == ".xlsx":
        from openpyxl import load_workbook

        ws = load_workbook(caminho, data_only=True).worksheets[0]
        it = ws.iter_rows(values_only=True)
        next(it, None)
        return [r for r in it if r and r[0]]

    for enc in ("utf-8-sig", "latin-1"):
        try:
            with caminho.open(encoding=enc, newline="") as fh:
                rs = list(csv.reader(fh, delimiter="\t", quotechar='"'))
        except UnicodeDecodeError:
            continue
        if rs and len(rs[0]) >= 4:
            return [tuple(r) for r in rs[1:]]
    return []


def carregar(pasta: Path | None = None) -> dict[str, dict[str, str]]:
    """ISSN normalizado -> {ciclo: estrato}, só da área de Computação."""
    pasta = pasta or DADOS
    out: dict[str, dict[str, str]] = {}
    for caminho in sorted(pasta.glob(PADRAO)):
        m = re.search(r"(\d{4}-\d{4})", caminho.name)
        if not m or m.group(1) not in CICLOS:
            continue
        ciclo = m.group(1)
        for r in _linhas(caminho):
            if len(r) < 4 or "COMPUTA" not in str(r[2]).upper():
                continue
            issn, estrato = normalizar_issn(r[0]), str(r[3] or "").strip()
            if len(issn) == 8 and estrato:
                out.setdefault(issn, {})[ciclo] = estrato
    return out


def buscar(issns: list[str], base: dict[str, dict[str, str]]) -> list[list[str]]:
    """[[ciclo, estrato], ...] em ordem cronológica, para um periódico.

    Um periódico costuma ter ISSN impresso e eletrônico, e a CAPES às vezes
    classificou os dois — em ciclos diferentes, inclusive. Unimos, e se os dois
    aparecerem no MESMO ciclo com estratos diferentes fica o melhor: é o que a
    comissão concedeu àquele periódico em alguma de suas formas.
    """
    achados: dict[str, str] = {}
    for i in issns or []:
        for ciclo, estrato in base.get(normalizar_issn(i), {}).items():
            atual = achados.get(ciclo)
            if atual is None or _ordem(estrato) < _ordem(atual):
                achados[ciclo] = estrato
    return [[c, achados[c]] for c in CICLOS if c in achados]


def _ordem(estrato: str) -> tuple[int, int]:
    """Menor é melhor. A antes de B antes de C; dentro da letra, número menor."""
    m = re.match(r"([ABC])(\d*)", estrato.strip().upper())
    if not m:
        return (9, 9)
    return ("ABC".index(m.group(1)), int(m.group(2) or 0))
