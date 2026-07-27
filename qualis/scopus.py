"""Leitura do Scopus Source List oficial — o percentil que a CAPES realmente usa.

A Elsevier publica gratuitamente a planilha "Scopus Source List", que traz, por
periódico e por categoria ASJC, o **CiteScore Percentile**. Esse é exatamente o
indicador do Procedimento 2 para periódicos (junto com o JIF Percentile do WoS,
valendo o maior dos dois).

Baixe a planilha em https://www.elsevier.com/products/scopus/content
(seção "Scopus source list") e aponte para ela:

    python -m qualis periodico "Journal of the ACM" --scopus-xlsx ~/Downloads/ext_list.xlsx

Requer `openpyxl` (veja requirements.txt).
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

# Categorias ASJC de Computação (1700-1712) e a genérica de Ciência da Computação.
ASJC_COMPUTACAO = tuple(str(c) for c in range(1700, 1713))


@dataclass
class LinhaScopus:
    nome: str
    issns: list[str]
    citescore: float | None
    percentil: float | None
    categoria: str | None


def _norm(s: object) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def _norm_issn(s: object) -> str:
    return re.sub(r"[^0-9xX]", "", str(s or "")).upper()


def _achar_colunas(header: list[str]) -> dict[str, int]:
    """Localiza as colunas de interesse tolerando variação de nome entre edições."""
    idx: dict[str, int] = {}
    for i, h in enumerate(header):
        h_low = h.lower()
        if "percentile" in h_low and "percentile" not in idx:
            idx["percentile"] = i
        elif "citescore" in h_low and "citescore" not in idx:
            idx["citescore"] = i
        elif h_low.startswith("source title") or h_low == "title":
            idx.setdefault("title", i)
        elif "print-issn" in h_low or h_low == "issn":
            idx.setdefault("issn", i)
        elif "e-issn" in h_low or "eissn" in h_low:
            idx.setdefault("eissn", i)
        elif "asjc" in h_low or "scopus sub-subject area" in h_low:
            idx.setdefault("categoria", i)
    return idx


def carregar(caminho: str | Path) -> list[LinhaScopus]:
    """Lê a planilha e retorna as linhas com percentil disponível.

    A planilha tem várias abas; procuramos a primeira com uma coluna de
    percentil e uma de título.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise ImportError(
            "openpyxl é necessário para ler o Scopus Source List: pip install openpyxl"
        ) from e

    wb = load_workbook(caminho, read_only=True, data_only=True)
    for aba in wb.worksheets:
        linhas = aba.iter_rows(values_only=True)
        try:
            header = [_norm(c) for c in next(linhas)]
        except StopIteration:
            continue
        idx = _achar_colunas(header)
        if "percentile" not in idx or "title" not in idx:
            continue

        out: list[LinhaScopus] = []
        for row in linhas:
            if not row or idx["title"] >= len(row):
                continue
            nome = _norm(row[idx["title"]])
            if not nome:
                continue
            issns = [
                _norm_issn(row[idx[k]])
                for k in ("issn", "eissn")
                if k in idx and idx[k] < len(row)
            ]
            out.append(
                LinhaScopus(
                    nome=nome,
                    issns=[i for i in issns if i],
                    citescore=_float(row, idx.get("citescore")),
                    percentil=_float(row, idx.get("percentile")),
                    categoria=(
                        _norm(row[idx["categoria"]])
                        if "categoria" in idx and idx["categoria"] < len(row)
                        else None
                    ),
                )
            )
        if out:
            return out

    raise ValueError(
        f"Não achei uma aba com colunas de título e percentil em {caminho}. "
        "Confira se é mesmo o Scopus Source List."
    )


def _float(row: tuple, i: int | None) -> float | None:
    if i is None or i >= len(row):
        return None
    try:
        return float(str(row[i]).replace(",", "."))
    except (TypeError, ValueError):
        return None


def melhor_percentil(
    linhas: list[LinhaScopus], consulta: str, *, so_computacao: bool = True
) -> tuple[float, LinhaScopus] | None:
    """Maior percentil do periódico entre suas categorias.

    A CAPES usa o maior percentil disponível; quando `so_computacao`, restringe
    às categorias ASJC de Computação (1700-1712), que é a leitura fiel para a
    Área 02.
    """
    q = consulta.strip().lower()
    issn_q = _norm_issn(consulta)
    candidatas = [
        l
        for l in linhas
        if l.percentil is not None
        and (l.nome.lower() == q or q in l.nome.lower() or (issn_q and issn_q in l.issns))
    ]
    if so_computacao:
        cs = [
            l
            for l in candidatas
            if l.categoria and any(a in l.categoria for a in ASJC_COMPUTACAO)
        ]
        candidatas = cs or candidatas
    if not candidatas:
        return None
    melhor = max(candidatas, key=lambda l: l.percentil or 0.0)
    return float(melhor.percentil), melhor
